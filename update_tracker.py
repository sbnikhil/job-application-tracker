import sys, json, openpyxl
from openpyxl.styles import PatternFill
from openpyxl.worksheet.datavalidation import DataValidation

path = '/Users/' + __import__('os').getenv('USER') + '/Desktop/JobRadar/job-tracker.xlsx'
rows = json.loads(sys.argv[1])

wb = openpyxl.load_workbook(path)
ws = wb['Applications']

# Status colors
fills = {
    'Applied':    PatternFill("solid", fgColor="D9EAF7"),
    'Assessment': PatternFill("solid", fgColor="FFF3CD"),
    'Interview':  PatternFill("solid", fgColor="D4EDDA"),
    'Rejected':   PatternFill("solid", fgColor="F8D7DA"),
    'Offer':      PatternFill("solid", fgColor="C3E6CB"),
}

# Clear existing data rows
for row in ws.iter_rows(min_row=2):
    for cell in row:
        cell.value = None
        cell.fill = PatternFill()

# Write new rows
headers = ['Company','Role','Status','Date Applied','Last Updated','Sender','Notes']
for i, row in enumerate(rows, start=2):
    for j, key in enumerate(headers, start=1):
        cell = ws.cell(row=i, column=j, value=row.get(key, ''))
        if key == 'Status' and row.get(key) in fills:
            for col in range(1, 8):
                ws.cell(row=i, column=col).fill = fills[row.get(key)]

# Ensure dropdown still exists
dv = DataValidation(type="list", formula1='"Applied,Assessment,Interview,Rejected,Offer"', allow_blank=True, showDropDown=False)
dv.sqref = "C2:C1000"
ws.add_data_validation(dv)

wb.save(path)
print("saved")
